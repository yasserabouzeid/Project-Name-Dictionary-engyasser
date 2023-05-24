from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os
import openpyxl

# define a function for user prompt of excel sheets
def file_prompt(msg):
    # Prompt user to select Excel file
    root = Tk()
    root.withdraw()
    ini_file_path = os.getcwd()
    excel_file_path = askopenfilename(title=msg, filetypes=[("Excel files", "*.xls*")], initialdir=ini_file_path)
    root.destroy()
    return excel_file_path

file_path = file_prompt('Choose your matching excel file')
wb1 = openpyxl.load_workbook(file_path,read_only=True,data_only=True)
ws_match = wb1['Mataching']

file_path = file_prompt('Choose your Revenue excel file')
wb2 = openpyxl.load_workbook(file_path,read_only=True,data_only=True)
ws_rev = wb2['Revise Cost 2023'] #3rd sheet in this workbook always

file_path = file_prompt('Choose your POC excel file')
wb3 = openpyxl.load_workbook(file_path,read_only=True,data_only=True)
ws_poc = wb3['2022'] #only one sheet

# Creating a dictionary with our names from a predefined excel sheet that can be updated if more projects are added for the future
dict = {}
c = 0
for i in range(2,503): # assumed number of possible project names wouldn't exceed 500
    if ws_match.cell(i,2).value is not None:
        # key would be cost control naming and value would be financial's naming
        dict[ws_match.cell(i,5).value] = ws_match.cell(i,2).value
    else:
        c += 1
    if c > 3:
        break # to avoid iterating over empty rows

# defining 2nd disctionary contaiaing POC values and finance names
dict2 = {}
crr_cell = 0
c = 2
while crr_cell is not None:
    crr_cell = ws_poc.cell(2,c).value
    dict2[crr_cell] = [ws_poc.cell(10,c).value, ws_poc.cell(20,c).value]
    c+=1

# getting project names from revenue excel sheet
c=6
dict3 = {}
while ws_rev.cell(c,7).value is not None:
    cost_name = ws_rev.cell(c,7).value
    finance_name = dict.get(ws_rev.cell(c,7).value,0)
    if finance_name != 0:
        try:
            dict3[cost_name] = [dict2[finance_name][0], dict2[finance_name][1]]
        except:
            pass
    else:
        dict3[cost_name] = ['No value', 'No Value']
    c+=1

# outputting to csv
with open('Results.csv', 'w') as f:
    header = ['Project', 'Total Cost', 'Total Revenue']
    f.write(','.join(header) + '\n')
    for key in dict3.keys():
        row = [key, str(dict3[key][0]), str(dict3[key][1])]
        f.write(','.join(row) + '\n')

wb1.close()
wb2.close()
wb3.close()
